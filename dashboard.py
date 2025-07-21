# dashboard.py
import pandas as pd
import logging
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.cell.cell import Cell
from pathlib import Path
import numpy as np
from datetime import datetime

log_file = f"Logs/Dashboard01.log"
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    handlers=[
        logging.FileHandler(log_file),
        logging.StreamHandler()
    ]
)
def standardize_columns(df):
    """Lowercase and strip all column names for consistency."""
    df.columns = [str(col).strip().lower() for col in df.columns]
    return df

def safe_index_to_str(series):
    s = series.copy()
    s.index = s.index.map(str)
    return s.groupby(s.index).sum()

def load_and_standardize(path, dtype, column_map=None):
    df = pd.read_csv(path, dtype=dtype, low_memory=False)
    df = standardize_columns(df)
    if column_map:
        df.rename(columns={k.lower(): v for k, v in column_map.items()}, inplace=True)
    # Always ensure a 'date' column
    if "date" not in df.columns or df["date"].isnull().all():
        # Try to extract from known columns
        for col in ["call_start_time", "data_requested_time", "sms_date"]:
            if col in df.columns and not df[col].isnull().all():
                df["date"] = pd.to_datetime(df[col], errors="coerce").dt.date
                break
        # If still missing, extract from folder name
        if "date" not in df.columns or df["date"].isnull().all():
            date_str = Path(path).parent.name
            try:
                date_val = pd.to_datetime(date_str, errors="coerce").date()
                df["date"] = date_val
            except Exception:
                df["date"] = pd.NaT
    return df

def load_all_inputs(base_dir):
    logging.info("Loading all input files...")
   
    call_dtypes = {
        'customer_id': 'str', 'imsi': 'str', 'calling_party': 'str', 'called_party': 'str',
        'billing_type': 'str', 'service_type': 'str', 'call_direction': 'str', 'call_category': 'str',
        "call_start_time": 'str', "call_end_time": 'str', 
        "available_limit": 'float', "deducted_limit": 'float',"available_onn_calls": 'int', 
        "consumed_onn_calls": 'int',"available_ofn_calls": 'int',"consumed_ofn_calls": 'int',
        "customer_balance": 'float',"current_consumed_balance":'float',"total_consumed_balance":'float'
    }
    sms_dtypes = {
        'customer_id': 'str', 'imsi': 'str', 'msisdn': 'str', "sms_receiver": 'str', 'sms_request': 'int', 'billing_type': 'str', 
        'service_type': 'str', 'call_direction': 'str', 'call_category': 'str','date': 'str', 
        "available_limit": 'float', "deducted_limit": 'float',"available_onn_sms": 'int', 
        "consumed_onn_sms": 'int',"available_ofn_sms": 'int',"consumed_ofn_sms": 'int'
    }
    data_dtypes = {
        'customer_id': 'str', 'imsi': 'str',  'msisdn': 'str', 'billing_type': 'str', 'service_type': 'str', 
        'call_direction': 'str', 'call_category': 'str', 'data_requested_time': 'str', 
        'total_consumed_data': 'int', "available_limit": 'float', "deducted_limit": 'float', 
        "available_data": 'int', "consumed_data": 'int', 'consumed_request': 'int'
    }

    dfs = {
        "calls": pd.concat([
            load_and_standardize(base_dir / "postpaid_call.csv", call_dtypes),
            load_and_standardize(base_dir / "prepaid_call.csv", call_dtypes)
        ], ignore_index=True),
        "sms": pd.concat([
            load_and_standardize(base_dir / "postpaid_sms.csv", sms_dtypes),
            load_and_standardize(base_dir / "prepaid_sms.csv", sms_dtypes)
        ], ignore_index=True),
        "data": pd.concat([
            load_and_standardize(base_dir / "postpaid_data.csv", data_dtypes),
            load_and_standardize(base_dir / "prepaid_data.csv", data_dtypes)
        ], ignore_index=True),
        "sim_inventory": standardize_columns(pd.read_csv(base_dir / "sim_inventory_export2.csv", dayfirst=True, parse_dates=["allocation_date"])),
        "activations": standardize_columns(pd.read_csv(base_dir / "ActivatedSIMcustomer.csv", parse_dates=["create_date"])),
        "partners": standardize_columns(pd.read_csv(base_dir / "partner.csv")),
    }
    
    # Convert sim_inventory and activations dates
    if "allocation_date" in dfs["sim_inventory"].columns:
        dfs["sim_inventory"]["allocation_date"] = pd.to_datetime(dfs["sim_inventory"]["allocation_date"], errors="coerce", dayfirst=True).dt.date
    if "create_date" in dfs["activations"].columns:
        dfs["activations"]["create_date"] = pd.to_datetime(dfs["activations"]["create_date"], errors="coerce", dayfirst=True).dt.date
    logging.info("Finished loading files.")
    return dfs

def get_user_column(df):
    for col in ["msisdn", "customer_id", "imsi"]:
        if col in df.columns:
            return df[["date", col]].dropna()
    return df[["date"]].assign(dummy_user="unknown")

def flexible_category_match(series, patterns):
    """Return a boolean Series matching any of the given patterns (case-insensitive, flexible)."""
    regex = "|".join([fr"\\b{p.replace('_', '[_-]?')}\\b" for p in patterns])
    return series.astype(str).str.lower().str.contains(regex, na=False, regex=True)


def summarize_sim(df_inventory, df_activation):
    import pandas as pd
    from datetime import date
    from dateutil.relativedelta import relativedelta

    logging.info("Summarizing SIM data...")
    summaries = {}

    def summarize_with_monthly(df, date_col, value_name):
        df = df.copy()
        df[date_col] = pd.to_datetime(df[date_col], errors="coerce").dt.date
        df = df.dropna(subset=[date_col])
        if df.empty:
            logging.info(f"No valid dates found for {value_name}. Skipping.")
            return pd.DataFrame()

        today = date.today()
        current_month = today.replace(day=1)

        # Daily (current month)
        full_days = pd.date_range(start=current_month, end=today).date
        df_current = df[df[date_col] >= current_month]
        daily_counts = df_current.groupby(date_col).size().reindex(full_days, fill_value=0)
        daily_cols = [d.strftime("%d-%m-%Y") for d in daily_counts.index]
        daily_vals = list(daily_counts)[::-1]
        daily_cols = daily_cols[::-1]

        # Monthly (prior months), with full range padded
        df_past = df[df[date_col] < current_month]
        if not df_past.empty:
            df_past["month"] = pd.Series(df_past[date_col]).astype("datetime64[ns]").dt.to_period("M").dt.to_timestamp()
            monthly_counts = df_past.groupby("month").size()

            # Create full range from earliest to current-1
            start_month = monthly_counts.index.min()
            end_month = current_month - relativedelta(months=1)
            full_months = pd.date_range(start=start_month, end=end_month, freq='MS')
            monthly_counts = monthly_counts.reindex(full_months, fill_value=0)

            full_months_desc = full_months[::-1]
            month_names = [d.strftime("%b'%y") for d in full_months_desc]  # Descending order
            monthly_vals = list(monthly_counts.reindex(full_months_desc, fill_value=0).values)

        else:
            month_names = []
            monthly_vals = []

        all_cols = daily_cols + month_names
        all_vals = daily_vals + monthly_vals

        summary = pd.DataFrame([all_vals], columns=all_cols)
        summary["Grand Total"] = summary.sum(axis=1)
        summary = summary[["Grand Total"] + all_cols]
        summary.index = [value_name]
        return summary

    # SIMs Sold
    if "allocation_date" in df_inventory.columns:
        sold_summary = summarize_with_monthly(df_inventory, "allocation_date", "SIMs Sold")
        if not sold_summary.empty:
            summaries["SIMs Sold"] = sold_summary
            logging.info("✔ SIMs Sold summary added.")
        else:
            logging.info("⚠ No data for SIMs Sold.")

    # SIMs Activated (prepaid / postpaid)
    if "billing_type" not in df_inventory.columns:
        df_inventory["billing_type"] = "Prepaid"

    if "customertype" in df_activation.columns:
        for billing_type in ["prepaid", "postpaid"]:
            df_act = df_activation[df_activation["customertype"].str.lower() == billing_type]
            if df_act.empty:
                logging.info(f"⚠ No activation data for {billing_type}.")
                continue

            act_summary = summarize_with_monthly(df_act, "create_date", f"{billing_type.capitalize()} SIMs Activated")
            if not act_summary.empty:
                summaries[billing_type] = act_summary
                logging.info(f"✔ {billing_type.capitalize()} SIMs Activated summary added.")
            else:
                logging.info(f"⚠ No rows for {billing_type} after summarization.")

    logging.info(f"✅ Finished SIM summary with {len(summaries)} sections.")
    return summaries


def summarize_usage(df_calls, df_sms, df_data):
    logging.info("Summarizing usage data...")
    summaries = {}
    for billing_type in ["prepaid", "postpaid"]:
        # Filter for current billing type
        calls = df_calls[df_calls["billing_type"] == billing_type].copy() if "billing_type" in df_calls.columns else pd.DataFrame()
        sms = df_sms[df_sms["billing_type"] == billing_type].copy() if "billing_type" in df_sms.columns else pd.DataFrame()
        data = df_data[df_data["billing_type"] == billing_type].copy() if "billing_type" in df_data.columns else pd.DataFrame()
                  
        # Parse date columns
        if "call_start_time" in calls.columns:
            calls["date"] = pd.to_datetime(calls["call_start_time"], errors="coerce").dt.date
        
        # Try to create a 'date' column from the most likely date columns
        if "date" in sms.columns:
            sms["date"] = pd.to_datetime(sms["date"], errors="coerce").dt.date
            
        if "data_requested_time" in data.columns:
            data["data_requested_time"] = pd.to_datetime(data["data_requested_time"], errors="coerce").dt.date
        
        # Filter onnet/offnet outgoing calls/sms
        onnet_outgoing = calls[calls["subcategory"].str.contains("onnet", na=False)] if "subcategory" in calls.columns else pd.DataFrame()
        offnet_outgoing = calls[calls["subcategory"].str.contains("offnet", na=False)] if "subcategory" in calls.columns else pd.DataFrame()
        
        sms_onnet = sms[sms["subcategory"].astype(str).str.contains("onnet", na=False)] if "subcategory" in sms.columns else pd.DataFrame()
        sms_offnet = sms[sms["subcategory"].astype(str).str.contains("offnet", na=False)] if "subcategory" in sms.columns else pd.DataFrame()
        
        
        # Aggregate metrics (safe for empty DataFrames)
        if "consumed_onn_calls" in onnet_outgoing.columns and "consumed_ofn_calls" in offnet_outgoing.columns and not calls.empty:
            voice_outgoing = (onnet_outgoing.groupby("date")["consumed_onn_calls"].sum() + offnet_outgoing.groupby("date")["consumed_ofn_calls"].sum()) / 60
        else:
            voice_outgoing = pd.Series(dtype=float)
        onnet_calls = onnet_outgoing.groupby("date")["consumed_onn_calls"].sum() / 60 if "consumed_onn_calls" in onnet_outgoing.columns and not onnet_outgoing.empty else pd.Series(dtype=float)
        offnet_calls = offnet_outgoing.groupby("date")["consumed_ofn_calls"].sum() / 60 if "consumed_ofn_calls" in offnet_outgoing.columns and not offnet_outgoing.empty else pd.Series(dtype=float)
        
        
        if "sms_request" in sms.columns and not sms.empty:
            sms_outgoing = sms.groupby("date")["sms_request"].sum()
        else:
            sms_outgoing = pd.Series(dtype=float)
        onn_sms = sms_onnet.groupby("date")["sms_request"].sum() if "sms_request" in sms_onnet.columns and not sms_onnet.empty else pd.Series(dtype=float)
        ofn_sms = sms_offnet.groupby("date")["sms_request"].sum() if "sms_request" in sms_offnet.columns and not sms_offnet.empty else pd.Series(dtype=float)
        
        
        if "consumed_data" in data.columns:
            prepaid_data = data.groupby("date")["consumed_data"].sum() / 1024**3 if "consumed_data" in data.columns and not data.empty else pd.Series(dtype=float)
        if "consumed_request" in data.columns:
            postpaid_data = data.groupby("date")["consumed_request"].sum() / 1024**3 if "consumed_request" in data.columns and not data.empty else pd.Series(dtype=float)
        
        # Active users
        user_calls = get_user_column(calls) if not calls.empty else pd.DataFrame()
        user_sms = get_user_column(sms) if not sms.empty else pd.DataFrame()
        user_data = get_user_column(data) if not data.empty else pd.DataFrame()
        active_users = pd.concat([user_calls, user_sms, user_data]).drop_duplicates().groupby("date").size() if not (user_calls.empty and user_sms.empty and user_data.empty) else pd.Series(dtype=int)
        
        if billing_type == "prepaid":
            total_data = prepaid_data
        elif billing_type == "postpaid":
            total_data = postpaid_data
        
        summary = pd.DataFrame({
            "Voice [Outgoing Mins]": safe_index_to_str(voice_outgoing),
            "On-net [Outgoing]": safe_index_to_str(onnet_calls),
            "Off-net [Outgoing]": safe_index_to_str(offnet_calls),
            "SMS [Outgoing Count]": safe_index_to_str(sms_outgoing),
            "on-net [SMS]" :safe_index_to_str(onn_sms),
            "off-net [SMS]" :safe_index_to_str(ofn_sms),
            "Data Usage[GB]" :safe_index_to_str(total_data),
            "Active Users": safe_index_to_str(active_users),
        }).fillna(0).T
        
        if summary.empty:
            continue
        summary["Grand Total"] = summary.sum(axis=1)
        summary.columns = [str(c) for c in summary.columns]  # Ensure all columns are strings
        summary = summary[["Grand Total"] + sorted([c for c in summary.columns if c != "Grand Total"], reverse=True)]
        
        summaries[billing_type] = summary
    return summaries

def summarize_partner_activations(df_activations, df_partners):
    import pandas as pd
    from datetime import date
    from dateutil.relativedelta import relativedelta

    logging.info("Summarizing partner SIM activations...")
    summaries = {}

    today = date.today()
    current_month = today.replace(day=1)
    full_days = pd.date_range(start=current_month, end=today).date
    full_day_strs = [d.strftime("%d-%m-%Y") for d in full_days][::-1]  # Newest to oldest

    for billing_type in ["prepaid", "postpaid"]:
        df = df_activations[df_activations["customertype"].str.lower() == billing_type].copy()
        if df.empty:
            logging.info(f"No partner activation data for {billing_type}.")
            continue

        df["partner_id"] = df["partner_id"].astype(str)
        df["create_date"] = pd.to_datetime(df["create_date"], errors="coerce", dayfirst=True).dt.date
        df_partners["id"] = df_partners["id"].astype(str)
        merged = df.merge(df_partners, left_on="partner_id", right_on="id", how="left")
        merged["business_name"] = merged.get("business_name", "Unknown Partner")

        # Add 'month' and 'day' cols
        merged["month"] = pd.to_datetime(merged["create_date"]).dt.to_period("M").dt.to_timestamp()
        merged["day"] = pd.to_datetime(merged["create_date"]).dt.strftime("%d-%m-%Y")

        # Separate current and past
        current = merged[merged["create_date"] >= current_month]
        past = merged[merged["create_date"] < current_month]

        # --- Daily (current month)
        daily = current.groupby(["business_name", "day"]).size().unstack(fill_value=0)
        for day in full_day_strs:
            if day not in daily.columns:
                daily[day] = 0
        daily = daily[full_day_strs]  # Order columns

        # --- Monthly (previous months)
        if not past.empty:
            past["month"] = pd.to_datetime(past["create_date"]).dt.to_period("M").dt.to_timestamp()
            monthly = past.groupby(["business_name", "month"]).size().unstack(fill_value=0)

            # Pad all months
            min_month = monthly.columns.min()
            max_month = current_month - relativedelta(months=1)
            full_months = pd.date_range(start=min_month, end=max_month, freq="MS")[::-1]  # Descending
            month_labels = [d.strftime("%b'%y") for d in full_months]

            for ts, label in zip(full_months, month_labels):
                if ts not in monthly.columns:
                    monthly[ts] = 0
            monthly = monthly[full_months]
            monthly.columns = month_labels
        else:
            monthly = pd.DataFrame()

        # --- Combine
        combined = pd.concat([daily, monthly], axis=1).fillna(0).astype(int)
        combined["Grand Total"] = combined.sum(axis=1)
        combined = combined[["Grand Total"] + [c for c in combined.columns if c != "Grand Total"]]
        combined = combined.loc[~(combined == 0).all(axis=1)]  # Drop zero rows
        summaries[billing_type] = combined

    return summaries


def generate_historical_summaries(dfs):
    import datetime
    logging.info("Generating historical summaries...")
    result = {}
    usage_dict = summarize_usage(dfs["calls"], dfs["sms"], dfs["data"])
    current_year = datetime.datetime.now().year
    for billing_type in ["prepaid", "postpaid"]:
        if billing_type not in usage_dict:
            continue
        usage = usage_dict[billing_type].T.copy()
        usage.index = pd.to_datetime(usage.index, errors="coerce", format="%Y-%m-%d")
        usage = usage.dropna()
        if usage.empty:
            continue
        usage_current_year = usage[usage.index.year == current_year]
        def fmt(df, freq, fmt_str):
            df = df.resample(freq).sum()
            df["sort_key"] = df.index
            df.index = df.index.strftime(fmt_str)
            df = df.sort_values("sort_key", ascending=False).drop(columns="sort_key")
            return df.T
        result[f"{billing_type} - Weekly Summary"] = fmt(usage_current_year, "W-MON", "W%W(%d%b)")
        result[f"{billing_type} - Monthly Summary"] = fmt(usage, "ME", "%b %Y")
        result[f"{billing_type} - Yearly Summary"] = fmt(usage, "YE", "%Y")
    return result

def write_table(wb, sheet_name, content_dict):
    logging.info(f"Writing sheet: {sheet_name}")
    ws = wb.create_sheet(sheet_name)
    row_offset = 1
    for section, df in content_dict.items():
        ws.merge_cells(start_row=row_offset, start_column=1, end_row=row_offset, end_column=df.shape[1] + 1)
        cell = ws.cell(row=row_offset, column=1, value=section)
        cell.fill = PatternFill(start_color="FF4958A6", end_color="FF4958A6", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True, size=8)
        row_offset += 1
        _write_df(ws, df, row_offset)
        row_offset += len(df) + 2
    _auto_width(ws)

def _write_df(ws, df, row_offset):
    df.index.name = None
    rows = list(dataframe_to_rows(df, index=True, header=True))
    if len(rows) > 0 and all(x is None for x in rows[0]):
        rows = rows[1:]
    for r_idx, row in enumerate(rows):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=row_offset + r_idx, column=c_idx, value=int(round(val)) if isinstance(val, float) else val)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(size=8)
            if r_idx == 0:
                cell.fill = PatternFill(start_color="FF4958A6", end_color="FF4958A6", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True, size=8)

def _auto_width(ws):
    fixed_width = 8.6
    for i, col in enumerate(ws.columns, start=1):
        col_letter = get_column_letter(i)
        values = [cell.value for cell in col if isinstance(cell, Cell) and cell.value]
        if values:
            max_len = max(len(str(v)) for v in values)
            if i == 1:
                ws.column_dimensions[col_letter].width = min(max_len + 2, 60)
            else:
                ws.column_dimensions[col_letter].width = fixed_width

def build_dashboard(dfs, out_path):
    logging.info("Building dashboard...")
    wb = Workbook()
    wb.remove(wb.active)
    write_table(wb, "Sim Summary", summarize_sim(dfs["sim_inventory"], dfs["activations"]))
    write_table(wb, "Usage Summary", summarize_usage(dfs["calls"], dfs["sms"], dfs["data"]))
    write_table(wb, "Partner SIM Activations", summarize_partner_activations(dfs["activations"], dfs["partners"]))
    write_table(wb, "Performance Summary", generate_historical_summaries(dfs))
    wb.save(out_path)
    logging.info(f"Dashboard saved to {out_path}")

if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--base", default="./Files", help="Path to folder containing input files")
    parser.add_argument("--out", default="Dashboard01.xlsx", help="Path to output Excel file")
    args = parser.parse_args()
    base_dir = Path(args.base)
    dfs = load_all_inputs(base_dir)
    build_dashboard(dfs, args.out)
